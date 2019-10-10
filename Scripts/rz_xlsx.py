"""
This is a script for manipulating xlsx
files. It includes  functions  to  add
modify, delete  and  get  data from an
xlsx file.
No relational  or  database  logic  is
implemented here.


To do: add meta data for tables and columns
to be used for user interaction and doc.
"""

MODEL_PATH = '../Model/'

from   openpyxl            import Workbook, load_workbook
from   openpyxl.utils.cell import get_column_letter
import sys

def get_model_sheet(table_name):
    """
    Centralizes the fetching of the Excel file main
    sheet of the object 'table_name'.
    Returns the active sheet of the table.

    Input:
    - table_name: name of the queried table/object

    Output:
    The active sheet of the Excel file representing
    the object.
    """
    
    table = load_workbook(MODEL_PATH+table_name)
    return table.active
    

def get_table_content(table_name,mode):
    """
    Returns table content by column or by row.
    mode = 'C' corresponds to the column mode
    whereby the returned dictionary has column
    names as keys, and column content as values.
    In the mode = 'R', row mode, we get a dict-
    ionary with the row number as the key, and
    the row content as the value. Row number 0
    is always assigned to the column names.
    Returns None if neither R nor C is used as
    a mode.

    Input:
    - table_name: name of the queried table
    - mode: data format (by row or by column)
    """
    table_dict = {}
    table = load_workbook(MODEL_PATH+table_name)
    sheet = table.active

    if mode == 'C': #column mode
        for column in list(sheet.columns):
            table_dict[str(column[0].value)] = []
            for i in range(1,len(column)):
                table_dict[str(column[0].value)].append(str(column[i].value))
    elif mode == 'R': #row mode
        rownum = 0
        for row in list(sheet.rows):
            table_dict[rownum] = []
            for i in range(len(row)):
                table_dict[rownum].append(str(row[i].value))
            rownum+=1
    else:
        print('Unknown data extraction mode.')
        return None
    
    return table_dict

def get_row_by_id(table_name,row_id):
    """
    Returns the row identified by its ID. If the row
    is not found, returns -1.

    Input:
    - table_name: name of the queried table

    Output:
    A list representing the desired row, or -1 if
    such a row is not found or an error occurs.
    """
    
    sheet     = get_model_sheet(table_name)
    id_column = [id_num.value for id_num in list(sheet.columns)[0]]
    #print(id_column)  #for debugging
    try:
        return [cell.value for cell in list(sheet.rows)[id_column.index(row_id)]]
    except:
        print(sys.exc_info())
        return -1

    
def get_table_headers(table_name):
    table_headers = []
    return table_headers

def get_id_by_search_key(table_name,column_name,search_key):
    id_list      = []
    sheet        = get_table_sheet(table_name)
    column_names = [col_name.value for col_name in list(sheet.rows)[0]]
    id_column    = [id_num.value for id_num in list(sheet.columns)[0]]
    try:
        column       = list(sheet.columns)[column_names.index(column_name)]
    except:
        print('No such column name')
    
    for i in range(len(column)):
        if column[i].value == search_key:
            id_list.append(id_column[i])
    
    return id_list

def create_row_with_user_input(table_name):
    sheet        = get_model_sheet(table_name)
    column_names = [col_name.value for col_name in list(sheet.rows)[0]][1:]
    id_column    = [int(id_num.value) for id_num in list(sheet.columns)[0][1:]]

    max_id       = max(id_column)
    
    input_values = [max_id+1]
    for label in column_names:
        input_values.append(input('Enter '+label+': '))

    return input_values
    

def add_row(table_name,row):
    """
    Adds a new row 'row' to the table 'table_name', and
    returns True if the operation is successful.
    If an exception occurs, prints the error message
    and returns False.

    Input:
    - table_name: name of the table (stored as an Excel
    file
    - row: a list of potentially heterogenous values
    representing the contents of the new row of the table.

    Output:
    A boolean denoting the success of the operation or
    lack thereof.
    """
    wb        = load_workbook(MODEL_PATH+table_name)
    sheet     = wb.active
    id_column = [id_num.value for id_num in list(sheet.columns)[0]]

    fill_row_num  = len(id_column)+1

    #get_column_letter

    try:
        for j in range(len(row)):
            print(get_column_letter(j+1)+str(fill_row_num))
            sheet[get_column_letter(j+1)+str(fill_row_num)] = row[j]
        wb.save(MODEL_PATH+table_name)    
        return True
    except:
        print('add_row failed')
        print(sys.exc_info())
        return False

def delete_row(table_name,row_id):
    return False

def modify_row(table_name,new_row):
    """
    Modifies a row identified by the ID new_row[0] by
    replacing its values with the values of 'new_row'
    in the table 'table_name', and returns True if the
    operation is successful.
    If an exception occurs, prints the error message
    and returns False.

    Input:
    - table_name: name of the table (stored as an Excel
    file
    - new_row: a list of potentially heterogenous values
    representing the contents of the new row to replace
    the old row of the table.

    Output:
    A boolean denoting the success of the operation or
    lack thereof.
    """
    wb        = load_workbook(MODEL_PATH+table_name)
    sheet     = wb.active
    id_column = [id_num.value for id_num in list(sheet.columns)[0]]
    
    row_index = id_column.index(new_row[0])+1

    #get_column_letter

    try:
        for j in range(len(new_row)):
            print(get_column_letter(j+1)+str(row_index))
            sheet[get_column_letter(j+1)+str(row_index)] = new_row[j]
        wb.save(MODEL_PATH+table_name)    
        return True
    except:
        print('modify_row failed')
        print(sys.exc_info())
        return False


if __name__ == '__main__':
    #some tests
    #table_name = 'activities.xlsx'
    table_name = 'objectives.xlsx'
    row_id     = 10
    #print(get_table_content(table_name,'R'))
    #print(get_row_by_id(table_name,row_id))

    #row = create_row_with_user_input(table_name)
    row = [7, 'Github 50', 'Have at least 50 repositories on Github', '', '31/12/2019', '2', 'False']
    print(modify_row(table_name,row))
